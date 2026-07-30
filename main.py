"""
API de collecte — Étude clinique ANALOG.

Deux liens, deux endpoints :
  POST /api/submit/medecin   -> une soumission médecin (T0 "avant" ou T1 "après", champ timepoint)
  POST /api/submit/patient   -> une soumission patient (14 items + pré-questionnaire)
  GET  /api/export?key=ADMIN_KEY -> génère et renvoie un .xlsx à jour (temps réel : lit direct la DB)
  GET  /api/health           -> ping

Données anonymes (pas de nom, pas d'identifiant patient) -> pas d'exigence HDS
sur ce backend. Le médecin renseigne lui-même un identifiant (M01...) pour
permettre l'appariement avant/après ; ce n'est pas une donnée patient.

Sécurité (pilote, ~10 médecins / ~500 patients) :
  - STUDY_KEY : clé partagée envoyée par le frontend à chaque soumission.
    Filtre anti-bot basique (visible dans le JS du formulaire), pas une
    authentification forte.
  - ADMIN_KEY : clé pour télécharger l'export -> à garder confidentielle.
"""
import io
import json
import os
import datetime
from typing import Optional

from fastapi import FastAPI, HTTPException, Query, Depends
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from starlette.responses import StreamingResponse

from database import SessionLocal, Submission, init_db
import scoring
import labels

STUDY_KEY = os.environ.get("STUDY_KEY", "changeme-study-key")
ADMIN_KEY = os.environ.get("ADMIN_KEY", "changeme-admin-key")
ALLOWED_ORIGINS = os.environ.get("ALLOWED_ORIGINS", "*").split(",")

VALID_TYPES = {"medecin", "patient"}

app = FastAPI(title="ANALOG Study API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_methods=["GET", "POST"],
    allow_headers=["*"],
)


@app.on_event("startup")
def on_startup():
    init_db()


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


class SubmissionPayload(BaseModel):
    study_key: str
    timepoint: Optional[str] = None          # T0 / T1 (médecin uniquement)
    medecin_id: Optional[str] = None         # auto-déclaré (médecin uniquement)
    centre: Optional[str] = None
    age: Optional[str] = None
    anciennete: Optional[str] = None         # médecin
    logiciel_actuel: Optional[str] = None    # médecin
    type_intervention: Optional[str] = None  # patient
    mode_prise_charge: Optional[str] = None  # patient
    support: Optional[str] = None            # patient
    answers: dict


@app.get("/api/health")
def health():
    return {"status": "ok", "time": datetime.datetime.utcnow().isoformat()}


@app.post("/api/submit/{questionnaire_type}")
def submit(questionnaire_type: str, payload: SubmissionPayload, db=Depends(get_db)):
    if questionnaire_type not in VALID_TYPES:
        raise HTTPException(404, f"Type de questionnaire inconnu: {questionnaire_type}")
    if payload.study_key != STUDY_KEY:
        raise HTTPException(403, "Clé d'étude invalide")

    # Normalisation de l'identifiant médecin : "m01", "M01 ", "M1" (espace/casse)
    # ne doivent pas casser l'appariement avant/après dans l'export.
    medecin_id_normalized = payload.medecin_id.strip().upper() if payload.medecin_id else None

    row = Submission(
        questionnaire_type=questionnaire_type,
        medecin_id=medecin_id_normalized,
        timepoint=payload.timepoint,
        centre=payload.centre,
        age=payload.age,
        anciennete=payload.anciennete,
        logiciel_actuel=payload.logiciel_actuel,
        type_intervention=payload.type_intervention,
        mode_prise_charge=payload.mode_prise_charge,
        support=payload.support,
        answers_json=json.dumps(payload.answers, ensure_ascii=False),
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return {"status": "recorded", "id": row.id}


# ============================== EXPORT EXCEL ==============================

HEADER_FILL = PatternFill(start_color="26215C", end_color="26215C", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center")
TITLE_FONT = Font(bold=True, size=14, color="26215C")
SUBTITLE_FONT = Font(italic=True, size=10, color="71717A")
KPI_LABEL_FONT = Font(size=10, color="71717A")
KPI_VALUE_FONT = Font(bold=True, size=18, color="26215C")


def _num(v):
    """Convertit en int/float si possible, pour que les cellules Excel soient de vrais nombres (SOMME, MOYENNE...)."""
    if v is None or v == "":
        return None
    try:
        if isinstance(v, str) and "." in v:
            return float(v)
        return int(v)
    except (TypeError, ValueError):
        return v


def _style_header_row(ws, row_idx, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN


def _add_table(ws, n_rows, n_cols, name):
    if n_rows < 1:
        return
    ref = f"A1:{get_column_letter(n_cols)}{n_rows + 1}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showRowStripes=True, showFirstColumn=False
    )
    ws.add_table(table)


def _autosize(ws, max_width=55, header_height=42):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, 12), max_width)
    ws.row_dimensions[1].height = header_height
    ws.freeze_panes = "A2"


def _kpi_card(ws, row, col, label_text, value):
    ws.cell(row=row, column=col, value=label_text).font = KPI_LABEL_FONT
    ws.cell(row=row + 1, column=col, value=value).font = KPI_VALUE_FONT


@app.get("/api/export")
def export(key: str = Query(...), db=Depends(get_db)):
    """
    Export temps réel : lit l'état actuel de la base à chaque appel, donc
    toujours à jour au moment du téléchargement.
    """
    if key != ADMIN_KEY:
        raise HTTPException(403, "Clé admin invalide")

    rows = db.query(Submission).order_by(Submission.submitted_at).all()
    medecin_rows = [r for r in rows if r.questionnaire_type == "medecin"]
    patient_rows = [r for r in rows if r.questionnaire_type == "patient"]
    medecin_avant = [r for r in medecin_rows if r.timepoint == "T0"]
    medecin_apres = [r for r in medecin_rows if r.timepoint == "T1"]

    wb = Workbook()
    wb.remove(wb.active)

    # ---------------------------------------------------------------
    # FEUILLE 1 — RÉSUMÉ (tableau de bord : 1 bloc médecin + 1 bloc patient)
    # ---------------------------------------------------------------
    ws = wb.create_sheet("Résumé")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Étude clinique ANALOG — Tableau de bord"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Généré le {datetime.datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws["A2"].font = SUBTITLE_FONT

    def _avg(vals):
        vals = [v for v in vals if v is not None]
        return round(sum(vals) / len(vals), 1) if vals else None

    def _delta(a, b):
        return round(b - a, 1) if a is not None and b is not None else None

    medecin_ids = {r.medecin_id for r in medecin_rows if r.medecin_id}

    sus_avant = _avg([scoring.sus_score(json.loads(r.answers_json), "av_") for r in medecin_avant])
    sus_apres = _avg([scoring.sus_score(json.loads(r.answers_json), "ap_") for r in medecin_apres])
    nasa_avant = _avg([scoring.nasa_tlx_score(json.loads(r.answers_json), "av_") for r in medecin_avant])
    nasa_apres = _avg([scoring.nasa_tlx_score(json.loads(r.answers_json), "ap_") for r in medecin_apres])
    bloc3_avant = _avg([scoring.bloc3_score(json.loads(r.answers_json), "av_") for r in medecin_avant])
    bloc3_apres = _avg([scoring.bloc3_score(json.loads(r.answers_json), "ap_") for r in medecin_apres])
    patient_scores = [scoring.patient_score(json.loads(r.answers_json))[0] for r in patient_rows]
    satisfaction_patient = _avg(patient_scores)

    # Couleurs par groupe logique (palette ANALOG déjà utilisée dans les questionnaires)
    C_ID = "26215C"       # violet ANALOG — comptages / identité
    C_SUS = "042C53"      # bleu — bloc SUS
    C_NASA = "633806"     # brun/amber — bloc NASA-TLX
    C_BLOC3 = "0F6E56"    # vert — bloc 3
    C_PATIENT = "9A3412"  # orange ANALOG — bloc patient

    def _styled_header(ws, row, col, text, hex_color):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        return cell

    # --- Bloc Médecins ---
    ws.cell(row=4, column=1, value="Médecins").font = Font(bold=True, size=12, color=C_ID)

    med_headers = [
        ("Nb réponses médecins (identifiants distincts)", C_ID),
        ("Nb questionnaires Avant ANALOG", C_ID),
        ("Nb questionnaires Après ANALOG", C_ID),
        ("SUS moyen — Avant (/100)", C_SUS),
        ("SUS moyen — Après (/100)", C_SUS),
        ("Différentiel SUS", C_SUS),
        ("NASA-TLX moyen — Avant (/50)", C_NASA),
        ("NASA-TLX moyen — Après (/50)", C_NASA),
        ("Différentiel NASA-TLX", C_NASA),
        ("Bloc 3 moyen — Avant (/100)", C_BLOC3),
        ("Bloc 3 moyen — Après (/100)", C_BLOC3),
        ("Différentiel Bloc 3", C_BLOC3),
    ]
    med_row = 5
    for i, (h, color) in enumerate(med_headers):
        _styled_header(ws, med_row, i + 1, h, color)
    ws.row_dimensions[med_row].height = 42

    med_values = [
        len(medecin_ids), len(medecin_avant), len(medecin_apres),
        sus_avant, sus_apres, _delta(sus_avant, sus_apres),
        nasa_avant, nasa_apres, _delta(nasa_avant, nasa_apres),
        bloc3_avant, bloc3_apres, _delta(bloc3_avant, bloc3_apres),
    ]
    for i, v in enumerate(med_values):
        ws.cell(row=med_row + 1, column=i + 1, value=v)

    ref = f"A{med_row}:{get_column_letter(len(med_headers))}{med_row + 1}"
    table = Table(displayName="ResumeMedecin", ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=False)
    ws.add_table(table)

    # --- Bloc Patients (ligne séparée, sous le bloc médecins) ---
    patient_section_row = med_row + 4
    ws.cell(row=patient_section_row - 1, column=1, value="Patients").font = Font(bold=True, size=12, color=C_PATIENT)

    pat_headers = [
        ("Nb réponses patients", C_PATIENT),
        ("Satisfaction patient moyenne (/100)", C_PATIENT),
    ]
    for i, (h, color) in enumerate(pat_headers):
        _styled_header(ws, patient_section_row, i + 1, h, color)
    ws.row_dimensions[patient_section_row].height = 42

    pat_values = [len(patient_rows), satisfaction_patient]
    for i, v in enumerate(pat_values):
        ws.cell(row=patient_section_row + 1, column=i + 1, value=v)

    ref = f"A{patient_section_row}:{get_column_letter(len(pat_headers))}{patient_section_row + 1}"
    table2 = Table(displayName="ResumePatient", ref=ref)
    table2.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=False)
    ws.add_table(table2)

    for col in range(1, len(med_headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 19

    # -----------------------------------------------------------
    # SECTION "INDICATEURS CLÉS" — synthèse chiffrée, prête pour un dossier/slide
    # -----------------------------------------------------------
    def _pct_change(before, after):
        """% de variation réelle avant->après (signe conservé). None si non calculable (donnée manquante ou avant=0)."""
        if before is None or after is None or before == 0:
            return None
        return round((after - before) / abs(before) * 100)

    def _sus_grade(score):
        if score is None:
            return None
        if score >= 80.3:
            return "Excellente"
        if score >= 68:
            return "Bonne"
        if score >= 51:
            return "Moyenne"
        return "Faible"

    def _fmt_pct(pct):
        if pct is None:
            return "n/d"
        return f"+{pct} %" if pct >= 0 else f"{pct} %"

    sus_pct = _pct_change(sus_avant, sus_apres)
    nasa_pct = _pct_change(nasa_avant, nasa_apres)  # négatif = charge réduite = amélioration
    bloc3_pct = _pct_change(bloc3_avant, bloc3_apres)

    def _to_int(v):
        try:
            return int(v)
        except (TypeError, ValueError):
            return None

    def _pct_ge(rows, key, threshold=4):
        """% de répondants ayant coché >= threshold (4 = 'plutôt d'accord' ou plus) sur UNE question précise."""
        vals = [_to_int(json.loads(r.answers_json).get(key)) for r in rows]
        vals = [v for v in vals if v is not None]
        if not vals:
            return None, 0
        return round(sum(1 for v in vals if v >= threshold) / len(vals) * 100), len(vals)

    def _pct_ge_avg(rows, keys, threshold=4):
        """% de répondants dont la MOYENNE de plusieurs questions est >= threshold."""
        vals = []
        for r in rows:
            a = json.loads(r.answers_json)
            nums = [_to_int(a.get(k)) for k in keys]
            nums = [n for n in nums if n is not None]
            if nums:
                vals.append(sum(nums) / len(nums))
        if not vals:
            return None, 0
        return round(sum(1 for v in vals if v >= threshold) / len(vals) * 100), len(vals)

    def _pct_patient_satisfait(rows, threshold=80):
        scores = [scoring.patient_score(json.loads(r.answers_json))[0] for r in rows]
        scores = [s for s in scores if s is not None]
        if not scores:
            return None, 0
        return round(sum(1 for s in scores if s >= threshold) / len(scores) * 100), len(scores)

    pct_recommande, n_recommande = _pct_ge(medecin_apres, "ap_c18")
    pct_continuite, n_continuite = _pct_ge(medecin_apres, "ap_c17")
    pct_adhesion, n_adhesion = _pct_ge_avg(medecin_apres, ["ap_c16", "ap_c17", "ap_c18"])
    pct_fiabilite, n_fiabilite = _pct_ge_avg(medecin_apres, ["ap_c13", "ap_c14", "ap_c15"])
    pct_facile, n_facile = _pct_ge(medecin_apres, "ap_s3")
    pct_confiant, n_confiant = _pct_ge(medecin_apres, "ap_s9")
    pct_patient_reco, n_patient_reco = _pct_ge(patient_rows, "p14")
    pct_patient_sat, n_patient_sat = _pct_patient_satisfait(patient_rows)

    headline_row = patient_section_row + 4
    ws.cell(row=headline_row, column=1, value="Chiffres clés — phrases pour présentation").font = Font(bold=True, size=15, color=C_ID)
    ws.cell(row=headline_row + 1, column=1,
            value="Basés sur des questions individuelles précises (pas des moyennes globales) — chaque chiffre indique sa source et sa taille d'échantillon (n). "
                  "Se recalcule automatiquement : plus l'échantillon grandit, plus ces pourcentages se consolident.")
    ws.cell(row=headline_row + 1, column=1).font = SUBTITLE_FONT
    ws.merge_cells(start_row=headline_row + 1, start_column=1, end_row=headline_row + 1, end_column=10)

    headline_bullets = []
    if pct_recommande is not None:
        headline_bullets.append((
            f"{pct_recommande} % des médecins recommanderaient ANALOG à un(e) collègue (Bloc 3, Q18 Après — n={n_recommande})",
            C_BLOC3,
        ))
    if pct_adhesion is not None:
        headline_bullets.append((
            f"{pct_adhesion} % des médecins adhèrent à ANALOG — satisfaits, comptent continuer à l'utiliser et le recommanderaient "
            f"(Bloc 3, Q16-17-18 Après — n={n_adhesion})",
            C_BLOC3,
        ))
    if pct_continuite is not None:
        headline_bullets.append((
            f"{pct_continuite} % des médecins prévoient de continuer à utiliser ANALOG régulièrement (Bloc 3, Q17 Après — n={n_continuite})",
            C_BLOC3,
        ))
    if pct_facile is not None:
        headline_bullets.append((
            f"{pct_facile} % des médecins jugent ANALOG facile à utiliser (SUS, Q3 Après — n={n_facile})",
            C_SUS,
        ))
    if pct_confiant is not None:
        headline_bullets.append((
            f"{pct_confiant} % des médecins se sentent confiants en utilisant ANALOG (SUS, Q9 Après — n={n_confiant})",
            C_SUS,
        ))
    if pct_fiabilite is not None:
        headline_bullets.append((
            f"{pct_fiabilite} % des médecins jugent ANALOG fiable et sécurisé pour leurs données patient (Bloc 3, Q13-15 Après — n={n_fiabilite})",
            C_BLOC3,
        ))
    if nasa_avant is not None and nasa_apres is not None:
        nasa_reduction = _pct_change(nasa_avant, nasa_apres)
        headline_bullets.append((
            f"La charge de travail perçue (mentale et physique) diminue de {abs(nasa_reduction)} % avec ANALOG "
            f"(NASA-TLX, {nasa_avant}/50 → {nasa_apres}/50 — n={len(medecin_avant)} avant / {len(medecin_apres)} après)",
            C_NASA,
        ))
    if pct_patient_sat is not None:
        headline_bullets.append((
            f"{pct_patient_sat} % des patients se déclarent satisfaits d'ANALOG (score ≥ 80/100 — n={n_patient_sat})",
            C_PATIENT,
        ))
    if pct_patient_reco is not None:
        headline_bullets.append((
            f"{pct_patient_reco} % des patients recommanderaient cette démarche à un proche (Q14 — n={n_patient_reco})",
            C_PATIENT,
        ))

    r = headline_row + 3
    for text, color in headline_bullets:
        cell = ws.cell(row=r, column=1, value="▶  " + text)
        cell.font = Font(bold=True, size=13, color=color)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
        ws.row_dimensions[r].height = 20
        r += 1

    if not headline_bullets:
        ws.cell(row=r, column=1, value="Pas encore assez de données pour générer ces chiffres — reviennent dès les premières réponses Après.")
        ws.cell(row=r, column=1).font = SUBTITLE_FONT
        r += 1

    r += 2

    # -----------------------------------------------------------
    # SECTION "INDICATEURS DÉTAILLÉS" — moyennes brutes avant/après (support des chiffres ci-dessus)
    # -----------------------------------------------------------
    insight_row = r
    ws.cell(row=insight_row, column=1, value="Indicateurs détaillés — moyennes avant/après").font = Font(bold=True, size=14, color=C_ID)
    ws.cell(row=insight_row + 1, column=1,
            value=f"Échantillon actuel : {len(medecin_avant)} questionnaire(s) Avant · {len(medecin_apres)} questionnaire(s) Après "
                  f"({len(medecin_ids)} médecin(s) distinct(s)) — ces pourcentages se consolident automatiquement à mesure que l'échantillon grandit.")
    ws.cell(row=insight_row + 1, column=1).font = SUBTITLE_FONT
    ws.merge_cells(start_row=insight_row + 1, start_column=1, end_row=insight_row + 1, end_column=8)

    bullets = []
    if sus_pct is not None:
        bullets.append((
            f"Utilisabilité perçue (SUS) : {_fmt_pct(sus_pct)}  —  {sus_avant}/100 → {sus_apres}/100 "
            f"(catégorie « {_sus_grade(sus_avant)} » → « {_sus_grade(sus_apres)} »)",
            C_SUS,
        ))
    if nasa_pct is not None:
        bullets.append((
            f"Charge de travail perçue (NASA-TLX) : {_fmt_pct(nasa_pct)}  —  {nasa_avant}/50 → {nasa_apres}/50 "
            f"(un pourcentage négatif = charge mentale/physique réduite)",
            C_NASA,
        ))
    if bloc3_pct is not None:
        bullets.append((
            f"Performance clinique globale (Bloc 3) : {_fmt_pct(bloc3_pct)}  —  {bloc3_avant}/100 → {bloc3_apres}/100",
            C_BLOC3,
        ))
    if satisfaction_patient is not None:
        bullets.append((
            f"Satisfaction patient moyenne : {str(satisfaction_patient).replace('.', ',')} % (n={len(patient_rows)} patient(s))",
            C_PATIENT,
        ))

    r = insight_row + 3
    for text, color in bullets:
        cell = ws.cell(row=r, column=1, value="●  " + text)
        cell.font = Font(bold=True, size=12, color=color)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        r += 1

    r += 2
    ws.cell(row=r, column=1,
            value="Cet onglet se recalcule automatiquement à chaque nouvel export — retélécharger le fichier pour une vue à jour.")
    ws.cell(row=r, column=1).font = SUBTITLE_FONT

    # ---------------------------------------------------------------
    # FEUILLE 2 — SCORES MÉDECIN (une ligne par soumission : T0 ou T1)
    # ---------------------------------------------------------------
    ws = wb.create_sheet("Scores médecin")

    C_IDENTITE = C_ID          # violet ANALOG — identité / contexte
    C_TXT = "444441"           # gris — retours qualitatifs neutres
    C_INCIDENT = "791F1F"      # rouge — colonnes sécurité/incident (attire l'œil dès l'en-tête)
    FILL_T0 = "EAF2FB"         # teinte de fond légère pour les lignes "Avant"
    FILL_T1 = "EAF7F0"         # teinte de fond légère pour les lignes "Après"
    FILL_OUI = "FBE2E2"        # surbrillance si un incident est déclaré "Oui"

    med_col_defs = [
        ("Identifiant", C_IDENTITE),
        ("Âge", C_IDENTITE),
        ("Centre", C_IDENTITE),
        ("Logiciel", C_IDENTITE),
        ("Timing (T0/T1)", C_IDENTITE),
        ("SUS total (/100)", C_SUS),
        ("NASA-TLX total (/50)", C_NASA),
        ("Bloc 3 total (/100)", C_BLOC3),
        ("Problèmes actuels — T0 (Q19 & Q20)", C_TXT),
        ("Dysfonctionnement — T1 (Q19)", C_INCIDENT),
        ("Situation à risque — T1 (Q20)", C_INCIDENT),
        ("Retour libre (Q21 T0 / Q22 T1)", C_TXT),
    ]
    for i, (h, color) in enumerate(med_col_defs):
        _styled_header(ws, 1, i + 1, h, color)
    ws.row_dimensions[1].height = 30

    def _oui_non(v):
        return {"oui": "Oui", "non": "Non"}.get(v, v)

    r = 2
    for row in medecin_rows:
        a = json.loads(row.answers_json)
        is_t0 = row.timepoint == "T0"
        prefix = "av_" if is_t0 else "ap_"

        sus_v = scoring.sus_score(a, prefix)
        nasa_v = scoring.nasa_tlx_score(a, prefix)
        bloc3_v = scoring.bloc3_score(a, prefix)

        problemes_t0 = None
        dysfonctionnement_t1 = None
        risque_t1 = None
        retour_libre = None

        if is_t0:
            q19 = a.get("av_txt_q19")
            q20 = a.get("av_txt_q20")
            parts = []
            if q19:
                parts.append(f"Difficultés : {q19}")
            if q20:
                parts.append(f"Fonctionnalités manquantes : {q20}")
            problemes_t0 = " | ".join(parts) if parts else None
            retour_libre = a.get("av_txt_q21")
        else:
            dysfonctionnement_t1 = _oui_non(a.get("ap_inc1"))
            risque_t1 = _oui_non(a.get("ap_inc2"))
            retour_libre = a.get("ap_txt_q22")

        values = [
            row.medecin_id, _num(row.age), row.centre, row.logiciel_actuel, row.timepoint,
            sus_v, nasa_v, bloc3_v,
            problemes_t0, dysfonctionnement_t1, risque_t1, retour_libre,
        ]
        row_fill = FILL_T0 if is_t0 else FILL_T1
        for i, v in enumerate(values):
            cell = ws.cell(row=r, column=i + 1, value=v)
            cell.fill = PatternFill(start_color=row_fill, end_color=row_fill, fill_type="solid")
        # Surbrillance rouge si un incident est déclaré "Oui" (sécurité — visible immédiatement)
        if dysfonctionnement_t1 == "Oui":
            ws.cell(row=r, column=10).fill = PatternFill(start_color=FILL_OUI, end_color=FILL_OUI, fill_type="solid")
            ws.cell(row=r, column=10).font = Font(bold=True, color=C_INCIDENT)
        if risque_t1 == "Oui":
            ws.cell(row=r, column=11).fill = PatternFill(start_color=FILL_OUI, end_color=FILL_OUI, fill_type="solid")
            ws.cell(row=r, column=11).font = Font(bold=True, color=C_INCIDENT)
        r += 1

    n_data_rows = r - 2
    if n_data_rows >= 1:
        ref = f"A1:{get_column_letter(len(med_col_defs))}{n_data_rows + 1}"
        table = Table(displayName="ScoresMedecin", ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=False)
        ws.add_table(table)
    _autosize(ws, max_width=45)

    # ---------------------------------------------------------------
    # FEUILLE 3 — SCORES PATIENT
    # ---------------------------------------------------------------
    ws = wb.create_sheet("Scores patient")
    headers = [
        "Centre", "Âge", "Type d'intervention", "Mode de prise en charge", "Support utilisé",
        "Score global (/100)", "Accès & espace patient (/100)", "Préconsultation (/100)",
        "Fonctionnement technique (/100)", "Impression globale (/100)", "Date de soumission",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    for r in patient_rows:
        answers = json.loads(r.answers_json)
        score, dims = scoring.patient_score(answers)
        ws.append([
            r.centre, _num(r.age), r.type_intervention, r.mode_prise_charge, r.support, score,
            dims["acces_espace_patient"], dims["preconsultation"], dims["technique"], dims["impression_globale"],
            r.submitted_at.strftime("%d/%m/%Y %H:%M") if r.submitted_at else None,
        ])
    _add_table(ws, len(patient_rows), len(headers), "ScoresPatient")
    _autosize(ws)

    # ---------------------------------------------------------------
    # FEUILLE 4 — RÉPONSES BRUTES MÉDECIN (en-têtes lisibles)
    # ---------------------------------------------------------------
    ws = wb.create_sheet("Réponses brutes médecin")
    all_keys = []
    for r in medecin_rows:
        for k in json.loads(r.answers_json).keys():
            if k not in all_keys:
                all_keys.append(k)
    meta_headers = ["ID médecin", "Moment", "Centre", "Âge", "Ancienneté (années)", "Logiciel actuel", "Date de soumission"]
    headers = meta_headers + [labels.label(k) for k in all_keys]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    for r in medecin_rows:
        answers = json.loads(r.answers_json)
        ws.append([
            r.medecin_id, {"T0": "Avant", "T1": "Après"}.get(r.timepoint, r.timepoint), r.centre,
            _num(r.age), _num(r.anciennete), r.logiciel_actuel,
            r.submitted_at.strftime("%d/%m/%Y %H:%M") if r.submitted_at else None,
        ] + [_num(answers.get(k)) if k not in ("inc1", "ap_inc1", "ap_inc2") else answers.get(k) for k in all_keys])
    _add_table(ws, len(medecin_rows), len(headers), "BrutesMedecin")
    _autosize(ws, max_width=70)

    # ---------------------------------------------------------------
    # FEUILLE 5 — RÉPONSES BRUTES PATIENT (en-têtes lisibles)
    # ---------------------------------------------------------------
    ws = wb.create_sheet("Réponses brutes patient")
    all_keys_p = []
    for r in patient_rows:
        for k in json.loads(r.answers_json).keys():
            if k not in all_keys_p:
                all_keys_p.append(k)
    meta_headers_p = ["Centre", "Âge", "Type d'intervention", "Mode de prise en charge", "Support", "Date de soumission"]
    headers = meta_headers_p + [labels.label(k) for k in all_keys_p]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    for r in patient_rows:
        answers = json.loads(r.answers_json)
        ws.append([
            r.centre, _num(r.age), r.type_intervention, r.mode_prise_charge, r.support,
            r.submitted_at.strftime("%d/%m/%Y %H:%M") if r.submitted_at else None,
        ] + [_num(answers.get(k)) if k != "support" else answers.get(k) for k in all_keys_p])
    _add_table(ws, len(patient_rows), len(headers), "BrutesPatient")
    _autosize(ws, max_width=70)

    wb.move_sheet("Résumé", offset=-len(wb.sheetnames))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"ANALOG_export_{datetime.date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
